from openpyxl import load_workbook
import re
from pprint import pprint
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.workbook.child import INVALID_TITLE_REGEX
import glob
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import tkinter.messagebox

# Print iterations progress
def printProgressBar (iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()

def main():
    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    while True:
        filename = askopenfilename(title="Select XLSX File") # show an "Open" dialog box and return the path to the selected file
        if(filename==""):
            if tkinter.messagebox.askretrycancel("Error",  "No XLSX file selected"):
                pass
            else:
                exit()
        else:
            if(filename.lower().endswith(".xlsx")):
                break
            else:
                if tkinter.messagebox.askretrycancel("Error",  "Selected file must be in XLSX format"):
                    pass
                else:
                    exit()

    path =  os.path.dirname(os.path.realpath(__file__))
    # folder_name = "source"
    # read_files = glob.glob(path+"/" + folder_name + "/*.xlsx")
    read_files = [filename]

    for read_file in read_files:
        # opening the source excel file
        workbook = load_workbook(read_file, data_only=True)
        sheet = workbook.worksheets[0]

        # delete all sheets except first sheet
        for title in workbook.sheetnames[1:]:
            del workbook[title]
        
        # set the start and end of row of data
        start = int(input("Enter the first column number of tenant data : "))
        end = int(input("Enter the last column number of tenant data : "))

        print("Getting Outstanding")
        printProgressBar(0, end-start, prefix = 'Progress:', suffix = 'Complete', length = 50)
        # loop trough row of data
        for i in range(start,end+1):
            # create new sheet with unit + tenant combined as sheet name
            sheet_name = str(sheet["C"+str(i)].value).lstrip(" ") + " " + str(sheet["B"+str(i)].value).lstrip(" ") 
            sheet_name = re.sub(INVALID_TITLE_REGEX, '_', sheet_name)   

            title = sheet_name if len(sheet_name) <= 31 else sheet_name[:31]
            workbook.create_sheet(title) # no more than 31 char
            # print(sheet_name)


            # write unit and name on created sheet
            created_sheet = workbook[title]
            created_sheet.column_dimensions['B'].width = 15
            created_sheet.column_dimensions['C'].width = 15
            created_sheet.column_dimensions['D'].width = 15
            created_sheet.column_dimensions['E'].width = 15
            created_sheet.column_dimensions['F'].width = 15

            created_sheet["B2"] = sheet["C"+str(i)].value # unit
            created_sheet["C2"] = sheet["B"+str(i)].value # tenant
            
            # border style
            thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='medium'))

            # set dict for all value per month
            data_dict = {}
            denda = {}

            # get utility value per month
            value = "Denda"
            denda = get_all_total_value(i, value, denda, sheet)

            # get utility value per month
            value = "Utility"
            data_dict = get_all_total_value(i, value, data_dict, sheet)
            # set table header
            created_sheet["C3"] = value
            created_sheet["C3"].border = thin_border
            created_sheet["C3"].alignment = Alignment(horizontal='center')
            created_sheet["C3"].font = Font(bold=True)


            # get utility value per month
            value = "Service Charge"
            data_dict = get_all_total_value(i, value, data_dict, sheet)
            # set table header
            created_sheet["D3"] = value
            created_sheet["D3"].border = thin_border
            created_sheet["D3"].alignment = Alignment(horizontal='center')
            created_sheet["D3"].font = Font(bold=True)
            
            # get utility value per month
            value = "Sinking Fund"
            data_dict = get_all_total_value(i, value, data_dict, sheet)
            # set table header
            created_sheet["E3"] = value
            created_sheet["E3"].border = thin_border
            created_sheet["E3"].alignment = Alignment(horizontal='center')
            created_sheet["E3"].font = Font(bold=True)

            # set table header total/bulan
            created_sheet["F3"] = "Total/Month"
            created_sheet["F3"].border = thin_border
            created_sheet["F3"].alignment = Alignment(horizontal='center')
            created_sheet["F3"].font = Font(bold=True)

            begin = 4
            total_ut = 0
            total_sc = 0
            total_sf = 0
            total_total_month = 0

            # Iterating over values
            for month, bill in data_dict.items():
                ut = bill["Utility"] if "Utility" in bill else "0"
                sc = bill["Service Charge"] if "Service Charge" in bill else "0"
                sf = bill["Sinking Fund"] if "Sinking Fund" in bill else "0"

                ut = ut.lstrip(" ").replace("-", "") if isinstance(ut, str) else ut
                sc = sc.lstrip(" ").replace("-", "") if isinstance(sc, str) else sc
                sf = sf.lstrip(" ").replace("-", "") if isinstance(sf, str) else sf

                ut = ut if ut != "" else "0"
                sc = sc if sc != "" else "0"
                sf = sf if sf != "" else "0"

                ut = ut if ut is not None else "0"
                sc = sc if sc is not None else "0"
                sf = sf if sf is not None else "0"

                total_month = int(ut)+int(sc)+int(sf)

                total_ut += int(ut)
                total_sc += int(sc)
                total_sf += int(sf)
                total_total_month += int(total_month)

                if(ut != "0" or sc != "0" or sf != "0"):
                    created_sheet["B"+str(begin)] = month
                    created_sheet["C"+str(begin)] = ut if ut != "0" else ""
                    created_sheet["D"+str(begin)] = sc if sc != "0" else ""
                    created_sheet["E"+str(begin)] = sf if sf != "0" else ""
                    created_sheet["F"+str(begin)] = total_month
                    
                    created_sheet["B"+str(begin)].border = thin_border
                    created_sheet["C"+str(begin)].border = thin_border
                    created_sheet["D"+str(begin)].border = thin_border
                    created_sheet["E"+str(begin)].border = thin_border
                    created_sheet["F"+str(begin)].border = thin_border

                    created_sheet["B"+str(begin)].number_format = 'MMMM y'
                    created_sheet["C"+str(begin)].number_format = 'Rp #,##'
                    created_sheet["D"+str(begin)].number_format = 'Rp #,##'
                    created_sheet["E"+str(begin)].number_format = 'Rp #,##'
                    created_sheet["F"+str(begin)].number_format = 'Rp #,##'

                    begin += 1
                    
            if total_total_month != 0:
                created_sheet["C"+str(begin)] = total_ut if total_ut != 0 else ""
                created_sheet["D"+str(begin)] = total_sc if total_sc != 0 else ""
                created_sheet["E"+str(begin)] = total_sf if total_sf != 0 else ""
                created_sheet["F"+str(begin)] = total_total_month

                created_sheet["C"+str(begin)].number_format = 'Rp #,##' if total_ut != 0 else ""
                created_sheet["D"+str(begin)].number_format = 'Rp #,##' if total_sc != 0 else ""
                created_sheet["E"+str(begin)].number_format = 'Rp #,##' if total_sf != 0 else ""
                created_sheet["F"+str(begin)].number_format = 'Rp #,##'


                denda = denda["Denda"] if denda["Denda"] is not None else 0
                created_sheet["E"+str(begin+1)] = "Late Charge"
                if denda != 0:
                    created_sheet["F"+str(begin+1)] = denda

                    created_sheet["F"+str(begin+1)].number_format = 'Rp #,##'
                else:
                    created_sheet["F"+str(begin+1)] = "-"

                created_sheet["E"+str(begin+1)].font = Font(bold=True)

                created_sheet["E"+str(begin+1)].alignment = Alignment(horizontal='right')
                created_sheet["F"+str(begin+1)].alignment = Alignment(horizontal='right')


                created_sheet["E"+str(begin+2)] = "Total Bill"
                created_sheet["F"+str(begin+2)] = total_ut + total_sc + total_sf + denda

                created_sheet["F"+str(begin+2)].number_format = 'Rp #,##'

                created_sheet["E"+str(begin+2)].font = Font(bold=True)
                created_sheet["F"+str(begin+2)].font = Font(bold=True)

                created_sheet["E"+str(begin+2)].alignment = Alignment(horizontal='right')
            else:
                del workbook[title]

            printProgressBar(i-start, end-start, prefix = 'Progress:', suffix = 'Complete', length = 50)        
        
        del workbook[workbook.sheetnames[0]]

        # checking dir
        if not os.path.exists(path + "/result"):
            # then create it.
            os.makedirs(path + "/result")
        # get csv file name
        print("Saving Outstanding...")
        file_name = read_file.split("\\")[-1].split("/")[-1][:-5]
        workbook.save(filename= path+"/" + "result" + "/Outstanding " + file_name + ".xlsx")
    print("DONE!")
    tkinter.messagebox.showinfo("Done", "Outstanding Table Generated")


def content_checker(sheet, value):
    for row in sheet:
        for cell in row:
            if cell.value == value:
                return cell

def merged_span_check(sheet, cell):
    for merged_cell in sheet.merged_cells.ranges:
        if cell.coordinate in merged_cell:
            return merged_cell

def get_all_total_value(i, value, data_dict, sheet):
    # get range value (ex. "A1:Z1") then splitted by ":"
    cell = content_checker(sheet, value)
    range_values = str(merged_span_check(sheet, cell)).split(":")

    # loop trough the range
    for r in sheet[range_values[0]:range_values[1]]:
        for x in r:
            # split letter and number using regex
            match = re.match(r"([a-z]+)([0-9]+)", x.coordinate, re.I)
            if match:
                # get splitted letter and number as an array
                coordinates = match.groups()

                month = sheet[coordinates[0]+str(int(coordinates[1])+1)].value
                os_val = sheet[coordinates[0]+str(i)].value

                # if the cell value is not empty
                if os_val is not None or os_val != "" or os_val != "0" or os_val != 0:
                    total = os_val
                else:
                    total = ""

                if value == "Denda":
                    data_dict["Denda"] = total
                else:
                    if month not in data_dict:
                        data_dict[month] = {}

                    data_dict[month][value] = total
    return data_dict

if __name__ == '__main__':
    main()